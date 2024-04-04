import openpyxl
import os
import pandas as pd
import json

os.system('clear') #Clear the terminal.

def custom_print(label, value):
    print(f"{label}: {value} \n")

def print_title(title):
    print(f"\n\n--------------------- {title}\n")

def print_dict_in_table(data):
    df = pd.DataFrame.from_dict(data, orient='index')
    print(df)


file_path = './PY02_EXCEL_WITH_PYTHON/company_revenue.xlsx'

wb = openpyxl.load_workbook(file_path)

sheets = wb.sheetnames
custom_print('sheets', sheets)
sheet1 = wb[sheets[0]]
cell_a1 = sheet1['A1']
custom_print('cell_a1', cell_a1)
custom_print('cell_a1 value', cell_a1.value)
custom_print('cell_a1 type', type(cell_a1))
custom_print('cell_a1 row', cell_a1.row)
custom_print('cell_a1 column', cell_a1.column)
custom_print('cell_a1 format', cell_a1.number_format)
custom_print('cell_a1 coordinate', cell_a1.coordinate)

print(f"{sheet1['A2'].value}, based in {sheet1['B2'].value} has a revenue of ${str(sheet1['c2'].value)} billion.")

print(f"Records x Fields of sheet1: {sheet1.max_row} x {sheet1.max_column}")

#Print Headers of Revenue table.
print_title('#Print Headers of Revenue table.')
max_col = sheet1.max_column
for i in range(1, max_col + 1):
    cell_obj = sheet1.cell(1, i)
    custom_print(cell_obj.coordinate, cell_obj.value)

#Print Records of Company column in Revenue table.
print_title('#Print Records of Company column in Revenue table.')
max_row = sheet1.max_row
for i in range(1,max_row+1):
    cell_obj = sheet1.cell(i,1)
    custom_print(cell_obj.coordinate, cell_obj.value)

#Get range of cells using iter_rows.
print_title('#Get range of cells using iter_rows.')
for row in sheet1.iter_rows(1, 2, 1, 3, True):
    print(row)

#Get range of cells using iter_cols.
print_title('#Get range of cells using iter_cols.')
for col in sheet1.iter_cols(1,2,1,max_row,True):
    print(col)

#Get range of data from table
print_title('#Get range of data from table.')
print(sheet1['A1':'C4'])


#Get range of data from table in tabular format.
revenues = {}
print_title('#Get range of data from table in tabular format.')
for row in sheet1.iter_rows(2, 4, 1, 3, True):
    company = row[0]
    revenue_details = {
        "Country": row[1],
        "Revenue": row[2]
    }
    revenues[company] = revenue_details
print_dict_in_table(revenues)

#Print revenues dictionary in json format.
print_title('#Print revenues dictionary in json format.')
print(json.dumps(revenues, indent=4))






print('\n------------------end of output------------------\n')
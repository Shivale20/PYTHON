# List of modules---------
import openpyxl
import os
import pandas as pd
import json

# List of variables---------
wb_file_path = '/Users/manishshivale/Documents/PYTHON/PY02_EXCEL_WITH_PYTHON/zomato_reviews.xlsx'
wb = openpyxl.load_workbook(wb_file_path)
sheet1 = wb['zomato']

# List of functions---------
def custom_print(label, value):
    print(f"{label}: {value} \n")

def print_title(title):
    print(f"\n\n--------------------- {title}\n")

def save_file():
    return wb.save(wb_file_path)

def print_dimension(sheet):
    custom_print('Sheet Dimension', sheet.calculate_dimension())

def print_headers_in_json(sheet):
    headers = {}
    max_col = sheet.max_column
    for i in range(1, max_col+1):
        cell_obj = sheet.cell(1,i)
        headers[cell_obj.coordinate] = cell_obj.value
    print(json.dumps(headers, indent=4))


#Clear the terminal.
os.system('clear') 

print_dimension(sheet1)

print_title('#Print Headers of table in json format.')
print_headers_in_json(sheet1)

print_title('#Print reviews of table in json format.')
REVIEWS = {}
max_cols = sheet1.max_column
for row in sheet1.iter_rows(2,5,1,9, True):
    ids = row[0]
    reviews_data = {
            "Restaurant Name" : row[1],
            "Country Code": row[2],
            "City": row[3],
            "Address": row[4],
            "Cuisines": row[5],
            "Price range": row[6],
            "Aggregate rating": row[7],
            "Rating text": row[8]
    }
    REVIEWS[ids] = reviews_data
print(json.dumps(REVIEWS, indent=4))

"""
sheetname.freeze_pane
This freezes columns left to given cell and rows above given cell.
"""
sheet1.freeze_panes = 'C2'
save_file()


